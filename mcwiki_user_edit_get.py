import json
import logging
import os
import threading
from copy import deepcopy
from functools import reduce
from time import strftime, localtime, sleep

import openpyxl as xl
from requests import get

__author__ = "QWERTY770"
__version__ = "2.0"
# 2023/11/09: MCW moved to https://zh.minecraft.wiki
# 2024/01/06: Updated the script to version 2.0, added multi-revision query to reduce the number of requests

logger = logging.getLogger('MCW EditCount Script')
logger.setLevel(logging.DEBUG)

fh = logging.FileHandler('editcount-script-v2.log', encoding="utf-8")
fh.setLevel(logging.DEBUG)
ch = logging.StreamHandler()
ch.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
fh.setFormatter(formatter)
ch.setFormatter(formatter)
logger.addHandler(fh)
logger.addHandler(ch)

# revisions api
rev_api = "https://zh.minecraft.wiki/api.php?action=query&format=json&prop=revisions&revids="
folder = os.getcwd()

namespace_names = {0: "（主）", 1: "讨论", 2: "用户", 3: "用户讨论",
                   4: "Minecraft Wiki", 5: "Minecraft Wiki讨论",
                   6: "文件", 7: "文件讨论", 8: "MediaWiki", 9: "MediaWiki讨论",
                   10: "模板", 11: "模板讨论", 12: "帮助", 13: "帮助讨论",
                   14: "分类", 15: "分类讨论", 274: "Widget", 275: "Widget talk",
                   420: "GeoJson", 421: "GeoJson talk", 828: "模块", 829: "模块讨论",
                   2300: "Gadget", 2301: "Gadget talk", 2302: "Gadget definition", 2303: "Gadget definition talk",
                   2900: "Map", 2901: "Map talk", 10000: "Minecraft Dungeons", 10001: "Minecraft Dungeons talk",
                   10002: "Minecraft Earth", 10003: "Minecraft Earth talk"}
namespace_names_keys = namespace_names.keys()
namespace_order = dict([(j, i) for i, j in enumerate(sorted(namespace_names.keys()))])


def get_page(url: str):
    try:
        req = get(url, timeout=8)
        if req.status_code != 200:
            logger.warning(f"Failed to get page {url}, code = {req.status_code}")
            sleep(5)
            return get_page(url)
        return req.text
    except Exception as err:
        logger.error(str(err))
        sleep(5)
        return get_page(url)


def get_revs(start, end):
    logger.debug(f"{start} to {end} started!")
    for i in range(start // 50, end // 50):
        data = get_page(rev_api + "%7C".join([str(rev) for rev in range(i * 50 + 1, i * 50 + 51)]))
        with open(os.path.join(os.path.join(folder, "rev"), f"rev_{i}.txt"), "w") as f:
            f.write(data)
    if end % 50 != 0:
        data = get_page(rev_api + "%7C".join([str(rev) for rev in range((end // 50) * 50 + 1, end + 1)]))
        with open(os.path.join(os.path.join(folder, "rev"), f"rev_{(end // 50)}.txt"), "w") as f:
            f.write(data)
    logger.debug(f"{start} to {end} finished!")


def get_edit_dic(start: int, end: int) -> dict:
    user_dic = {}
    for i in range(start // 50, end // 50):
        pth = os.path.join(os.path.join(folder, "rev"), f"rev_{i}.txt")
        if not os.path.exists(pth):
            get_revs(i, i)
        with open(pth, "r", encoding="utf-8") as f:
            js = json.loads(f.read())
        if "pages" not in js["query"]:
            continue
        pages = js["query"]["pages"]
        for p in pages:
            namespace = pages[p]["ns"]
            revs = pages[p]["revisions"]
            for rev in revs:
                try:
                    revid = rev["revid"]
                    if revid > end or revid < start:
                        continue
                    if "user" not in rev and "userhidden" in rev:
                        user = "<HIDDEN_USER>"
                        logger.debug(f"The user is hidden in revision {revid}")
                    else:
                        user = rev["user"]
                    if user not in user_dic:
                        user_dic[user] = {"all": 0}
                    if namespace not in user_dic[user]:
                        user_dic[user][namespace] = 1
                    else:
                        user_dic[user][namespace] += 1
                    user_dic[user]["all"] += 1
                except Exception as err:
                    logger.error(str(err))
                    logger.error(f"JSON={str(js)}")
                    continue
    return user_dic


def merge_edit_dic(dic1: dict, dic2: dict) -> dict:
    result = deepcopy(dic1)
    for username in dic1.keys():
        for namespace in dic1[username].keys():
            if username in dic2:
                if namespace in dic2[username]:
                    result[username][namespace] += dic2[username][namespace]
    del username, namespace
    for username in dic2.keys():
        for namespace in dic2[username].keys():
            if username not in result:
                result[username] = dic2[username]
            else:
                if namespace not in result[username]:
                    result[username][namespace] = dic2[username][namespace]
    return result


def make_workbook(dic: dict, filename=f"minecraftwiki-useredit-{strftime('%Y%m%d-%H%M%S', localtime())}.xlsx"):
    wb = xl.Workbook()
    ws = wb.create_sheet('main', 0)

    ws.cell(row=1, column=1).value = "用户名"
    ws.cell(row=1, column=2).value = "编辑总计"
    for a, b in enumerate(namespace_names.keys()):
        ws.cell(row=1, column=a + 3).value = namespace_names[b]

    for m, i in enumerate(dic.keys()):
        user = dic[i]
        ws.cell(row=m + 2, column=1).value = i
        ws.cell(row=m + 2, column=2).value = user["all"]
        for j in namespace_names_keys:
            if j in user:
                ws.cell(row=m + 2, column=namespace_order[j] + 3).value = user[j]
            else:
                ws.cell(row=m + 2, column=namespace_order[j] + 3).value = 0

    wb.save(os.path.join(folder, filename))
    wb.close()
    logger.info("Successfully generated workbook sheet(s)!")


def download_data():
    thread_list = []
    for i in range(32):
        t = threading.Thread(target=get_revs, args=(1 + 527 * 50 * i, 527 * 50 + 527 * 50 * i))
        t.start()
        thread_list.append(t)
    for j in thread_list:
        j.join()
    get_revs(843201, 844400)


def workbook():
    for i in range(168):
        with open(os.path.join(os.path.join(folder, "slices"), f"{i}.txt"), "w", encoding="utf-8") as f:
            f.write(str(get_edit_dic(1 + 5000 * i, 5000 + 5000 * i)))
    logger.info("Successfully generated 168 slices!")

    slices_list = []
    for i in range(168):
        with open(os.path.join(os.path.join(folder, "slices"), f"{i}.txt"), "r", encoding="utf-8") as f:
            slices_list.append(eval(f.read()))
    make_workbook(merge_edit_dic(reduce(merge_edit_dic, slices_list), get_edit_dic(840001, 844400)))


if __name__ == "__main__":
    # download_data()
    # workbook()
    logger.info("Finished!")
