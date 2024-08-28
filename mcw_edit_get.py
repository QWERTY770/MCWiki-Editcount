import json
import logging
import os
import threading
from copy import deepcopy
from functools import reduce
from time import strftime, localtime, sleep

import openpyxl as xl
import pywikiapi as wiki

__author__ = "QWERTY770"
__version__ = "3.4"

folder = os.getcwd()
logger = logging.getLogger('MCW EditCount Script')
logger.setLevel(logging.DEBUG)

if os.path.exists(os.path.join(folder, "config.json")):
    with open("config.json", "r", encoding="utf-8") as f:
        config = json.load(f)  # type: dict
        headers = config.setdefault("headers", {})
        username = config.setdefault("username", "")
        password = config.setdefault("password", "")
        per = config.setdefault("per_request", 50)
else:
    headers = {}
    username = password = ""
    per = 50

fh = logging.FileHandler('editcount-script-v3.log', encoding="utf-8")
fh.setLevel(logging.DEBUG)
ch = logging.StreamHandler()
ch.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
fh.setFormatter(formatter)
ch.setFormatter(formatter)
logger.addHandler(fh)
logger.addHandler(ch)

site = wiki.Site("https://zh.minecraft.wiki/api.php", retry_after_conn=15)
rev_api = "https://zh.minecraft.wiki/api.php?action=query&format=json&prop=revisions&revids="
namespace_names = {0: "（主）", 1: "讨论", 2: "用户", 3: "用户讨论",
                   4: "Minecraft Wiki", 5: "Minecraft Wiki讨论",
                   6: "文件", 7: "文件讨论", 8: "MediaWiki", 9: "MediaWiki讨论",
                   10: "模板", 11: "模板讨论", 12: "帮助", 13: "帮助讨论",
                   14: "分类", 15: "分类讨论", 828: "模块", 829: "模块讨论",
                   2300: "Gadget", 2301: "Gadget talk", 2302: "Gadget definition", 2303: "Gadget definition talk",
                   9996: "地下城教程", 9997: "地下城教程讨论", 9998: "教程", 9999: "教程讨论",
                   10000: "地下城", 10001: "地下城讨论", 10002: "地球", 10003: "地球讨论",
                   10004: "故事模式", 10005: "故事模式讨论", 10006: "传奇", 10007: "传奇讨论"}
namespace_names_keys = namespace_names.keys()
namespace_order = dict([(j, i) for i, j in enumerate(sorted(namespace_names.keys()))])

# variables
total_edits = 952800
threads = 4
per_thread = int(total_edits / per / threads)
total_slices = int(total_edits / 5000)


def get_rev(revids: str, index: int) -> None:
    try:
        data = site("query", prop="revisions", EXTRAS={"headers": headers}, revids=revids)
        file = open(os.path.join(folder, "rev", f"rev_{index}.txt"), "w", encoding="utf-8")
        json.dump(data, file, ensure_ascii=False)
    except Exception as err:
        logger.error(err)
        sleep(10)
        get_rev(revids, index)


def get_revs(start: int, end: int) -> None:
    logger.debug(f"{start} to {end} started!")
    for i in range(start // per, end // per):
        if i % 50 == 0:
            print(f"{start}-{end} {i}\n", end="")
        get_rev("|".join([str(rev) for rev in range(i * per + 1, i * per + 51)]), i)
    if end % per != 0:
        get_rev("|".join([str(rev) for rev in range((end // per) * per + 1, end + 1)]), end // per)
    logger.debug(f"{start} to {end} finished!")


def get_edit_dic(start: int, end: int) -> dict:
    if start > end:
        return {}
    user_dic = {}
    for i in range(start // per, end // per):
        pth = os.path.join(folder, "rev", f"rev_{i}.txt")
        if not os.path.exists(pth):
            get_revs(i, i)
        with open(pth, "r", encoding="utf-8") as file:
            content = file.read()
        while not content:
            get_revs(i * per + 1, i * per + per)
            with open(pth, "r", encoding="utf-8") as file:
                content = file.read()
        js = json.load(open(pth, "r", encoding="utf-8"))
        if "pages" not in js["query"]:
            continue
        pages = js["query"]["pages"]
        for p in pages:
            namespace = p["ns"]
            revs = p["revisions"]
            for rev in revs:
                try:
                    revid = rev["revid"]
                    if revid > end or revid < start:
                        continue
                    if "user" not in rev and "userhidden" in rev:
                        user = "<HIDDEN_USER>"
                        logger.debug(f"The user is hidden in revision {revid}")
                    else:
                        user = rev["user"]
                    if user not in user_dic:
                        user_dic[user] = {"all": 0}
                    if namespace not in user_dic[user]:
                        user_dic[user][namespace] = 1
                    else:
                        user_dic[user][namespace] += 1
                    user_dic[user]["all"] += 1
                except Exception as err:
                    logger.error(str(err))
                    logger.error(f"JSON={str(js)}")
                    continue
    return user_dic


def merge_edit_dic(dic1: dict, dic2: dict) -> dict:
    result = deepcopy(dic1)
    for username in dic1.keys():
        for namespace in dic1[username].keys():
            if username in dic2 and namespace in dic2[username]:
                result[username][namespace] += dic2[username][namespace]
    del username, namespace
    for username in dic2.keys():
        for namespace in dic2[username].keys():
            if username not in result:
                result[username] = dic2[username]
            elif namespace not in result[username]:
                result[username][namespace] = dic2[username][namespace]
    return result


def make_workbook(dic: dict, filename=None) -> None:
    if filename is None:
        filename = f"minecraftwiki-useredit-{strftime('%Y%m%d-%H%M%S', localtime())}.xlsx"
    wb = xl.Workbook()
    ws = wb.create_sheet("main", 0)
    ws.cell(row=1, column=1).value = "用户名"
    ws.cell(row=1, column=2).value = "编辑总计"
    for a, b in enumerate(namespace_names.keys()):
        ws.cell(row=1, column=a + 3).value = namespace_names[b]

    for m, i in enumerate(dic.keys()):
        user = dic[i]
        ws.cell(row=m + 2, column=1).value = i
        ws.cell(row=m + 2, column=2).value = user["all"]
        for j in namespace_names_keys:
            if j in user:
                ws.cell(row=m + 2, column=namespace_order[j] + 3).value = user[j]
            else:
                ws.cell(row=m + 2, column=namespace_order[j] + 3).value = 0

    wb.save(os.path.join(folder, filename))
    wb.close()


def download_data() -> None:
    if username and password:
        site.login(username, password)
    thread_list = []
    for i in range(threads):
        num = per_thread * per * i
        t = threading.Thread(target=get_revs, args=(num + 1, num + per_thread * per))
        t.start()
        thread_list.append(t)
    for i in thread_list:
        i.join()
    get_revs(per_thread * per * threads + 1, total_edits)


def workbook() -> None:
    for i in range(total_slices):
        slic = get_edit_dic(1 + 5000 * i, 5000 + 5000 * i)
        file = open(os.path.join(folder, "slices", f"{i}.json"), "w", encoding="utf-8")
        json.dump(slic, file, ensure_ascii=False, indent=2)
    logger.info(f"Successfully generated {total_slices} slices!")

    slices_list = []
    for i in range(total_slices):
        slic = json.load(open(os.path.join(folder, "slices", f"{i}.json"), "r", encoding="utf-8"))
        slic2 = {}
        for j in slic.keys():
            slic2[j] = {}
            for k in slic[j].keys():
                if k == "all":
                    slic2[j][k] = slic[j][k]
                else:
                    slic2[j][int(k)] = slic[j][k]
        slices_list.append(slic2)
    dic = merge_edit_dic(reduce(merge_edit_dic, slices_list), get_edit_dic(total_slices * 5000, total_edits))
    make_workbook(dic)
    logger.info("Successfully generated workbook sheet(s)!")


if __name__ == "__main__":
    download_data()
    workbook()
    logger.info("Finished!")
